from datetime import timedelta

import openpyxl
import pytz
from django.conf import settings as dj_settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Q, Prefetch
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.utils import timezone

from .models import Registro, LogAcao, LinkImportante, CategoriaApoio
from .forms import RegistroForm
from .filters import RegistroFilter, LogAcaoFilter


def usuario_pode_ver_dashboard(user):
    return user.is_staff or user.groups.filter(name="dashboard").exists()


def _diff_registro(old_obj, new_obj):
    """
    Retorna dict no formato:
    {"campo": {"old": "...", "new": "..."}, ...}
    incluindo tipo com nome antigo -> nome novo
    """
    changes = {}

    # campos texto
    for field in ["titulo", "texto"]:
        old_v = getattr(old_obj, field, "") or ""
        new_v = getattr(new_obj, field, "") or ""
        if old_v != new_v:
            changes[field] = {"old": old_v, "new": new_v}

    # tipo (FK) - salva NOME antigo -> NOME novo
    old_tipo_id = getattr(old_obj, "tipo_id", None)
    new_tipo_id = getattr(new_obj, "tipo_id", None)
    if old_tipo_id != new_tipo_id:
        old_nome = getattr(getattr(old_obj, "tipo", None), "nome", "") if old_tipo_id else ""
        new_nome = getattr(getattr(new_obj, "tipo", None), "nome", "") if new_tipo_id else ""
        changes["tipo"] = {"old": old_nome, "new": new_nome}

    return changes


def livro_home(request):
    qs = Registro.objects.select_related("tipo", "autor").all()

    f = RegistroFilter(request.GET, queryset=qs)
    registros = f.qs

    form = RegistroForm()

    if request.method == "POST":
        if not request.user.is_authenticated:
            return HttpResponseForbidden("Você precisa estar logado para adicionar registros.")

        form = RegistroForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.autor = request.user
            obj.save()

            # LOG CREATE (usuário real)
            LogAcao.objects.create(
                usuario=request.user,
                acao="CREATE",
                tipo=obj.tipo,
                registro_id=obj.id,
                titulo=obj.titulo or "",
                detalhes={}
            )

            messages.success(request, "Registro criado com sucesso.")
            return redirect("livro:home")

    context = {
        "filter": f,
        "registros": registros,
        "form": form,
    }
    return render(request, "livro_monitoramento/livro_home.html", context)


@login_required
def editar_registro(request, pk):
    registro = get_object_or_404(Registro, pk=pk)

    if not (request.user.is_staff or registro.autor_id == request.user.id):
        return HttpResponseForbidden("Sem permissão para editar.")

    if request.method == "POST":
        old = Registro.objects.select_related("tipo").get(pk=registro.pk)

        form = RegistroForm(request.POST, instance=registro)
        if form.is_valid():
            updated = form.save()

            changes = _diff_registro(old, updated)

            if changes:
                LogAcao.objects.create(
                    usuario=request.user,
                    acao="UPDATE",
                    tipo=updated.tipo,
                    registro_id=updated.id,
                    titulo=updated.titulo or "",
                    detalhes=changes
                )

            messages.success(request, "Registro atualizado.")
            return redirect("livro:home")
    else:
        form = RegistroForm(instance=registro)

    return render(request, "livro_monitoramento/editar_registro.html", {
        "form": form,
        "registro": registro
    })


@login_required
def excluir_registro(request, pk):
    registro = get_object_or_404(Registro, pk=pk)

    if not (request.user.is_staff or registro.autor_id == request.user.id):
        return HttpResponseForbidden("Sem permissão para excluir.")

    if request.method == "POST":
        # LOG DELETE antes de apagar
        LogAcao.objects.create(
            usuario=request.user,
            acao="DELETE",
            tipo=registro.tipo,
            registro_id=registro.id,
            titulo=registro.titulo or "",
            detalhes={}
        )

        registro.delete()
        messages.success(request, "Registro excluído.")
        return redirect("livro:home")

    return render(request, "livro_monitoramento/confirmar_exclusao.html", {"registro": registro})


def logs_home(request):
    qs = LogAcao.objects.select_related("usuario", "tipo").all()
    f = LogAcaoFilter(request.GET, queryset=qs)
    logs = f.qs

    return render(request, "livro_monitoramento/logs_home.html", {
        "filter": f,
        "logs": logs,
    })


def links_home(request):
    itens_ativos = LinkImportante.objects.filter(ativo=True).order_by("ordem", "titulo")

    categorias = (
        CategoriaApoio.objects
        .filter(ativo=True)
        .annotate(qtd=Count("itens", filter=Q(itens__ativo=True)))
        .filter(qtd__gt=0)
        .prefetch_related(Prefetch("itens", queryset=itens_ativos))
        .order_by("ordem", "nome")
    )

    sem_categoria = itens_ativos.filter(categoria__isnull=True)

    return render(request, "livro_monitoramento/links_home.html", {
        "categorias": categorias,
        "sem_categoria": sem_categoria,
    })


@login_required
def dashboard_home(request):
    if not usuario_pode_ver_dashboard(request.user):
        raise PermissionDenied("Você não tem permissão para acessar o dashboard.")

    hoje = timezone.localdate()
    inicio_7_dias = hoje - timedelta(days=6)

    registros = Registro.objects.select_related("tipo", "autor").all()

    total_registros = registros.count()
    registros_hoje = registros.filter(criado_em__date=hoje).count()

    por_tipo = (
        registros.values("tipo__nome")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    por_autor = (
        registros.values("autor__username")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    por_dia_qs = (
        registros.filter(criado_em__date__gte=inicio_7_dias, criado_em__date__lte=hoje)
        .values("criado_em__date")
        .annotate(total=Count("id"))
        .order_by("criado_em__date")
    )

    mapa_dias = {item["criado_em__date"]: item["total"] for item in por_dia_qs}
    serie_7_dias = []
    for i in range(7):
        dia = inicio_7_dias + timedelta(days=i)
        serie_7_dias.append({
            "dia": dia.strftime("%d/%m"),
            "total": mapa_dias.get(dia, 0),
            "total_px": mapa_dias.get(dia, 0) * 18,
        })

    ultimos_registros = registros.order_by("-criado_em")[:10]

    context = {
        "total_registros": total_registros,
        "registros_hoje": registros_hoje,
        "total_tipos": registros.values("tipo").distinct().count(),
        "por_tipo": list(por_tipo),
        "por_autor": list(por_autor),
        "serie_7_dias": serie_7_dias,
        "ultimos_registros": ultimos_registros,
    }
    return render(request, "livro_monitoramento/dashboard.html", context)


def exportar_excel(request):
    qs = Registro.objects.select_related("tipo", "autor").all()
    f = RegistroFilter(request.GET, queryset=qs)
    registros = f.qs

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Livro de Monitoramento"

    headers = ["#", "Tipo", "Título", "Descrição", "Autor", "Data/Hora"]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    tz = pytz.timezone(getattr(dj_settings, "TIME_ZONE", "UTC"))

    for row_idx, r in enumerate(registros, start=2):
        criado_local = r.criado_em.astimezone(tz).strftime("%d/%m/%Y %H:%M")
        autor_nome = r.autor.get_full_name() or r.autor.username
        ws.append([
            row_idx - 1,
            r.tipo.nome,
            r.titulo,
            r.texto,
            autor_nome,
            criado_local,
        ])
        ws.cell(row=row_idx, column=4).alignment = Alignment(wrap_text=True)

    col_widths = [6, 20, 40, 60, 25, 18]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="livro_monitoramento.xlsx"'
    wb.save(response)
    return response